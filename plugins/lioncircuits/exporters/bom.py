"""Bill of materials export, built straight from PCB footprint fields.

This reads whatever fields are present on each footprint (Value, Footprint,
and any custom fields such as MPN / Manufacturer / LCSC / Supplier /
Description). It works even without a linked schematic netlist, which is
what a PCB-editor action plugin has access to.
"""

import csv
import os
import re

# Custom fields we'll pull in if present, in this preferred column order.
KNOWN_FIELDS = ["MPN", "Manufacturer", "LCSC", "Supplier", "Description", "DNP"]

# Header text for LC's assembly-BOM upload field-mapper (see export_bom()).
# "Description" and "MPN" are handled directly in export_bom()'s fieldnames
# list since they map to LC's exact target labels; this covers the rest.
LC_FIELD_LABELS = {
    "MPN": "Manufacturer Part Number",
}


def _natural_key(ref):
    parts = re.split(r"(\d+)", ref)
    return [int(p) if p.isdigit() else p for p in parts]


def _get_field(fp, field_name):
    """Best-effort read of a footprint field across KiCad 6/7/8 API variants."""
    try:
        if fp.HasFieldByName(field_name):
            return fp.GetFieldByName(field_name).GetText()
    except AttributeError:
        pass
    try:
        props = fp.GetProperties()
        if field_name in props:
            return props[field_name]
    except AttributeError:
        pass
    return ""


def export_bom(board, out_dir, filename="bom_for_assembly.csv"):
    """Group footprints by (Value, Footprint, extra fields) and write a CSV
    BOM with one row per unique group and a combined Qty + References.

    Column headers match LionCircuits' assembly-BOM upload field names
    exactly (Item Number, Designator, Unit Quantity, Description,
    Manufacturer Part Number, Footprint, Value) so the site's auto-mapper
    can match them without the user having to manually map "Designator"
    or "Item Number" on every upload — those two didn't have any matching
    column before. Note: this depends on LC's site matching by header
    text, which isn't something verifiable from here without access to
    the site itself — worth confirming on your next upload that it
    actually auto-maps end to end.
    """
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, filename)

    groups = {}
    for fp in board.GetFootprints():
        # Skip footprints explicitly excluded from BOM (KiCad 7+).
        try:
            import pcbnew
            if fp.GetAttributes() & pcbnew.FP_EXCLUDE_FROM_BOM:
                continue
        except AttributeError:
            pass

        value = fp.GetValue()
        footprint_name = str(fp.GetFPID().GetLibItemName())
        extra = tuple((name, _get_field(fp, name)) for name in KNOWN_FIELDS)
        key = (value, footprint_name, extra)
        groups.setdefault(key, []).append(fp.GetReference())

    rows = []
    for (value, footprint_name, extra), refs in groups.items():
        refs_sorted = sorted(refs, key=_natural_key)
        row = {
            "Designator": ", ".join(refs_sorted),
            "Unit Quantity": len(refs_sorted),
            "Value": value,
            "Footprint": footprint_name,
        }
        for name, val in extra:
            if val:
                row[LC_FIELD_LABELS.get(name, name)] = val
        rows.append(row)

    rows.sort(key=lambda r: (r["Value"], r["Footprint"]))

    # "Item Number" is LC's own row-sequence field — assign after sorting
    # so it reflects the order the site will actually display.
    for i, row in enumerate(rows, start=1):
        row["Item Number"] = i

    fieldnames = (
        ["Item Number", "Designator", "Unit Quantity", "Description",
         "Manufacturer Part Number", "Footprint", "Value"]
        + [LC_FIELD_LABELS.get(name, name) for name in KNOWN_FIELDS
           if name not in ("Description", "MPN")]
    )
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

    return path


# --- Supplier quote-request BOM (XLSX) -------------------------------------
#
# Separate from export_bom() above: this produces the fab/sourcing intake
# format from the "BOMFORTEST" sample — Project/Currency/Number of Boards
# header block, a styled MPN/Description/Required Qty/Order Qty/Unit
# Price/Total Price/Alternates table, and a Subtotal/GST/Total block.
# It has no References column (the sample template doesn't have one) —
# use the CSV/HTML BOM above if you need per-reference traceability.

QUOTE_HEADER_FILL = "1F4E78"
QUOTE_TOTALS_FILL = "E7EEF7"
QUOTE_COLUMNS = ["MPN", "Description", "Required Qty", "Order Qty",
                 "Unit Price", "Total Price", "Alternates"]
QUOTE_COL_WIDTHS = {"A": 18, "B": 42, "C": 14, "D": 12, "E": 14, "F": 16, "G": 50}


def _quote_groups(board, number_of_boards):
    """Group footprints by MPN (falling back to Value when no MPN field is
    set) and return quote rows: (mpn, description, required_qty, alternates).
    """
    try:
        import pcbnew
    except ImportError:
        pcbnew = None

    groups = {}
    order = []
    for fp in board.GetFootprints():
        if pcbnew is not None:
            try:
                if fp.GetAttributes() & pcbnew.FP_EXCLUDE_FROM_BOM:
                    continue
            except AttributeError:
                pass

        mpn = _get_field(fp, "MPN")
        value = fp.GetValue()
        key = mpn if mpn else f"{value} (no MPN set)"
        if key not in groups:
            groups[key] = {
                "description": _get_field(fp, "Description") or value,
                "alternates": _get_field(fp, "Alternates"),
                "count": 0,
            }
            order.append(key)
        groups[key]["count"] += 1

    rows = []
    for key in order:
        g = groups[key]
        required_qty = g["count"] * number_of_boards
        rows.append((key, g["description"], required_qty, g["alternates"]))
    rows.sort(key=lambda r: r[0])
    return rows


def export_bom_quote(board, out_dir, project_name="", currency="INR",
                      number_of_boards=1, filename="bom_for_components.xlsx"):
    """Write the supplier quote-request BOM. Uses openpyxl if available for
    the full styled .xlsx; otherwise falls back to an equivalent .csv so
    the export never fails outright over a missing optional dependency.

    Returns the path written.
    """
    rows = _quote_groups(board, max(1, int(number_of_boards or 1)))

    try:
        import openpyxl
    except ImportError:
        return _write_quote_csv(rows, out_dir, project_name, currency,
                                 number_of_boards, filename)

    return _write_quote_xlsx(openpyxl, rows, out_dir, project_name, currency,
                              number_of_boards, filename)


def _write_quote_xlsx(openpyxl, rows, out_dir, project_name, currency,
                       number_of_boards, filename):
    from openpyxl.styles import Font, PatternFill

    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor=QUOTE_HEADER_FILL)
    totals_fill = PatternFill("solid", fgColor=QUOTE_TOTALS_FILL)

    ws["A1"] = "Project"; ws["A1"].font = bold; ws["B1"] = project_name
    ws["A2"] = "Currency"; ws["A2"].font = bold; ws["B2"] = currency
    ws["A3"] = "Number of Boards"; ws["A3"].font = bold; ws["B3"] = number_of_boards

    header_row = 5
    for i, col_name in enumerate(QUOTE_COLUMNS):
        cell = ws.cell(row=header_row, column=i + 1, value=col_name)
        cell.font = bold
        cell.fill = header_fill

    first_data_row = header_row + 1
    r = first_data_row
    for mpn, description, required_qty, alternates in rows:
        ws.cell(row=r, column=1, value=mpn)
        ws.cell(row=r, column=2, value=description)
        ws.cell(row=r, column=3, value=required_qty)
        ws.cell(row=r, column=4, value=required_qty)  # Order Qty defaults to Required Qty
        ws.cell(row=r, column=5, value=0)              # Unit Price — fill in from quote
        ws.cell(row=r, column=6, value=f"=D{r}*E{r}")  # Total Price
        ws.cell(row=r, column=7, value=alternates or "")
        r += 1
    last_data_row = r - 1

    totals_row = r + 1
    if last_data_row >= first_data_row:
        subtotal_formula = f"=SUM(F{first_data_row}:F{last_data_row})"
    else:
        subtotal_formula = 0
    for label, row_offset, value in (
        ("Subtotal", 0, subtotal_formula),
        ("GST", 1, 0),
        ("Total", 2, f"=F{totals_row}+F{totals_row + 1}"),
    ):
        row = totals_row + row_offset
        lbl_cell = ws.cell(row=row, column=5, value=label)
        lbl_cell.font = bold
        lbl_cell.fill = totals_fill
        val_cell = ws.cell(row=row, column=6, value=value)
        val_cell.fill = totals_fill

    for col, width in QUOTE_COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    wb.save(path)
    return path


def _write_quote_csv(rows, out_dir, project_name, currency, number_of_boards, filename):
    """Dependency-free fallback when openpyxl isn't installed in KiCad's
    Python environment. Same columns/data, plain CSV instead of styled xlsx.
    """
    os.makedirs(out_dir, exist_ok=True)
    csv_filename = os.path.splitext(filename)[0] + ".csv"
    path = os.path.join(out_dir, csv_filename)

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Project", project_name])
        writer.writerow(["Currency", currency])
        writer.writerow(["Number of Boards", number_of_boards])
        writer.writerow([])
        writer.writerow(QUOTE_COLUMNS)
        subtotal = 0
        for mpn, description, required_qty, alternates in rows:
            writer.writerow([mpn, description, required_qty, required_qty, 0, 0, alternates or ""])
        writer.writerow([])
        writer.writerow(["", "", "", "", "Subtotal", subtotal])
        writer.writerow(["", "", "", "", "GST", 0])
        writer.writerow(["", "", "", "", "Total", subtotal])

    return path
